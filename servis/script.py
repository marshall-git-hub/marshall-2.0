#python -m PyInstaller --onefile --icon=src/icon.ico src/script.py


#! =================================================================================
#! ==============================Krok 1 (Gather)====================================
#! =================================================================================
from pywinauto import Application, keyboard,mouse
import pyautogui
from pywinauto.keyboard import send_keys
import time
import cv2
import numpy as np
from PIL import ImageGrab

def capture_screenshot(region=None):
    # Capture a screenshot of the specified region
    screenshot = ImageGrab.grab(bbox=region)
    return screenshot

def find_image(template_path, screenshot):
    # Load the template image
    template = cv2.imread(template_path, cv2.IMREAD_GRAYSCALE)
    template_w, template_h = template.shape[::-1]
    
    # Convert screenshot to grayscale
    screenshot_np = np.array(screenshot)
    gray_screenshot = cv2.cvtColor(screenshot_np, cv2.COLOR_BGR2GRAY)
    
    # Perform template matching
    res = cv2.matchTemplate(gray_screenshot, template, cv2.TM_CCOEFF_NORMED)
    min_val, max_val, min_loc, max_loc = cv2.minMaxLoc(res)
    
    # Define a threshold to determine if the image is found
    threshold = 0.8
    if max_val >= threshold:
        return max_loc, (template_w, template_h)
    else:
        return None, None

def find_and_click(region, template_path, double = False):
    time.sleep(0.5)
    screenshot = capture_screenshot(region)
    position, size = find_image(template_path, screenshot)
    
    if position:
        x, y = position
        width, height = size
        if double == False:
            mouse.click(coords=(x + width // 2, y + height // 2))
       
region = (0, 0, 1920, 1080)  

def main():
    # Path to the executable
    app_path = r"C:\Doprava3K\InfosysMain.exe"

    # Start the application
    app = Application(backend="uia").start(app_path)
    time.sleep(2)  # Adjust delay as necessary for the application to fully load

    try:
        # Specify the main window
        window_title = "Přihlášení do databáze"
        main_window = app.window(title=window_title)
        
        # Input password
        password_field = main_window.child_window(control_type="Edit", found_index=0)
        password_field.type_keys("8344")  # Type the password

        # Click OK button
        ok_button = main_window.child_window(title="OK", control_type="Button")
        ok_button.click()

        try:
            # Wait for the first popup window to appear
            first_popup_title = "Doprava 4K"
            first_popup_window = app.window(title=first_popup_title)
            first_popup_window.wait("visible", timeout=20)  # Wait for the window to become visible

            # Click "Ne" (No) button in the first popup
            ne_button = first_popup_window.child_window(title="Ne", control_type="Button")
            ne_button.click()
            
        except:
            pass

        # Wait for the second popup window to appear
        second_popup_title = "Doprava 4K"
        second_popup_window = app.window(title=second_popup_title)
        second_popup_window.wait("visible", timeout=10)  # Wait for the window to become visible

        # Click "Ano" (Yes) button in the second popup
        ano_button = second_popup_window.child_window(title="Ano", control_type="Button")
        ano_button.click()


    
        time.sleep(3)  # Wait for 3 seconds
        keyboard.send_keys("{LEFT}")  # Press the left arrow key
        keyboard.send_keys("{ENTER}")  # Press Enter
    

        # Wait for the new window "Seznam vozidel s prošlou kontrolou" to appear after clicking "Ano"
        new_window_title = "Seznam vozidel s prošlou kontrolou"
        new_window = app.window(title=new_window_title)
        new_window.wait("visible", timeout=10)  # Adjust timeout as necessary
 
        find_and_click(region, r"src\excel_button_find.png")
       
        # Assuming you need to interact with the "Save As" window
        save_as_title = "Save As"
        save_as_window = app.window(title=save_as_title)
        save_as_window.wait("visible", timeout=10)  # Adjust timeout as necessary

        # Use the keyboard to focus on the address bar and type the path
        save_as_window.set_focus()
        keyboard.send_keys("^l")  # Ctrl + L to focus the address bar
        keyboard.send_keys(r"Z:\Jakub\scripts\KontrolaApp\src{ENTER}")

        time.sleep(1)  # Allow time for the address to be entered and directory to change

        # Find the "File name:" edit field and type "doprava"
        file_name_field = save_as_window.child_window(title="File name:", control_type="Edit")
        file_name_field.set_focus()
        file_name_field.type_keys("doprava")

        # Example: Press Enter to proceed with saving
        keyboard.send_keys("{ENTER}")

        time.sleep(1)
        find_and_click(region, r"src\find\ne_button.png")
        time.sleep(0.3)
        find_and_click(region, r"src\find\close_window_button.png")
        time.sleep(0.3)



        find_and_click(region, r"src\find\close_window_button.png")
        time.sleep(0.3)
        find_and_click(region, r"src\find\Doprava_button.png")
        time.sleep(0.3)
        find_and_click(region, r"src\find\Vozidla1_button.png")
        time.sleep(0.3)
        find_and_click(region, r"src\find\Vozidla2_button.png")
        
        time.sleep(5)
        find_and_click(region, r"src\excel_button_find.png")
        time.sleep(0.3)
       
        # Assuming you need to interact with the "Save As" window
        save_as_title = "Save As"
        save_as_window = app.window(title=save_as_title)
        save_as_window.wait("visible", timeout=10)  # Adjust timeout as necessary

        # Use the keyboard to focus on the address bar and type the path
        save_as_window.set_focus()
        keyboard.send_keys("^l")  # Ctrl + L to focus the address bar
        keyboard.send_keys(r"Z:\Jakub\scripts\KontrolaApp\src{ENTER}")

        time.sleep(1)  # Allow time for the address to be entered and directory to change

        # Find the "File name:" edit field and type "doprava"
        file_name_field = save_as_window.child_window(title="File name:", control_type="Edit")
        file_name_field.set_focus()
        file_name_field.type_keys("vozidla")

        # Example: Press Enter to proceed with saving
        keyboard.send_keys("{ENTER}")
        
    except Exception as e:
        print(f"Error: {str(e)}")
        # Handle the error accordingly, such as logging, retrying, or terminating the script
    finally:
        time.sleep(2)
        app.kill()


from openpyxl import load_workbook

def kontroly_gather():
    file_path = "Z:\\Jakub\\scripts\\KontrolaApp\\kontroly.xlsx"
    workbook = load_workbook(filename=file_path, data_only=True)
    sheet = workbook.active

    # List to store all tables' data
    rows_as_dicts = []

    # Define the starting columns of each table
    table_columns = [(2, 4), (6, 8), (10, 13)]  # Example: first table in columns B-D, second in F-H, etc.

    for min_col, max_col in table_columns:
        # Loop through each row in the specified range of columns for the current table
        for row in sheet.iter_rows(min_row=2, min_col=min_col, max_col=max_col, values_only=True):
            # Convert the row to a dictionary
            if min_col == 10:
                row_dict = {
                "SPZ": row[0],
                "DueTo": row[1],
                "Kontrola": row[2],
                "Note": row[3]
            }
            
            else:
                row_dict = {
                    "SPZ": row[0],
                    "DueTo": row[1],
                    "Note": row[2]
                }
            # Append the dictionary to the list

            rows_as_dicts.append(row_dict)

      
 
    return rows_as_dicts

def organize_data_by_category(data):
    # Initialize the nested dictionary with categories
    categories = {
        'STK + EK': [],
        'Tachograf': [],
        'DPF': [],
        'Ciachovanie': [],
        'L- Certifikát': [],
        
        'Ročná tahač': [],
        'Ročná náves': [],
        "Motor. olej": [],
        "Difer. olej": [],
        "Geometria": [],
        "Kontrola Bŕzd": [],

        "Ostatné": [],
        "Osobné": []
              
   }
    
    # Temporary variable to store the current category
    current_category = None

    for entry in data:
        spz = entry['SPZ']
        
        # Check if the SPZ is one of the category labels
        if spz in categories:
            current_category = spz
        elif current_category:
            # Append the entry to the current category list
            categories[current_category].append(entry)

    return categories

def tagra_gather():
    # Path to the executable
    executable_path = r"C:\Program Files (x86)\TAGRA.eu\appTAGRA.exe"
    region = (0, 0, 1920, 1080)  

    try:
        # Start the application
        app = Application(backend="uia").start(executable_path)

        # Get the main window
        main_window = app.window(title_re="TAGRA.eu")
        main_window.wait('visible', timeout=20)

        app = Application(backend="uia").connect(title_re=".*TAGRA\.eu.*")
        main_window = app.window(title_re=".*TAGRA\.eu.*")

        # Find and click "planovane_button_find"
        find_and_click(region, r"Z:\Jakub\scripts\KontrolaApp\src\planovane_button_find.png")
        time.sleep(0.5)  # Short delay to allow UI to respond

        # Find and click "tlac_button_find"
        find_and_click(region, r"Z:\Jakub\scripts\KontrolaApp\src\tlac_button_fing.png")
        time.sleep(0.5)  # Short delay to allow UI to respond

        # Find and click "tabulka_button_find"
        find_and_click(region, r"Z:\Jakub\scripts\KontrolaApp\src\tabulka_button_find.png")
        time.sleep(0.5)  # Short delay to allow UI to respond

        # Find and click "ulozit_button_find"
        find_and_click(region, r"Z:\Jakub\scripts\KontrolaApp\src\ulozit_button_find.png")
        time.sleep(0.5)  # Short delay to allow UI to respond

        

        keyboard.send_keys("^l")  # Ctrl + L to focus the address bar
        keyboard.send_keys(r"Z:\Jakub\scripts\KontrolaApp\src{ENTER}")

        time.sleep(1)  # Allow time for the address to be entered and directory to change

        # Find and click the "File name:" field using image recognition
        find_and_click(region, r"Z:\Jakub\scripts\KontrolaApp\src\file_name_find.png")

        # Type "tagra" into the field
        keyboard.send_keys("tagra")

        # Press Enter to proceed with saving
        keyboard.send_keys("{ENTER}")
        app.kill()

    except Exception as e:
        print(f"An error occurred: {e}")


if __name__ == "__main__":
    print("Starting Doprava4K.exe")
    main()
    #tagra_gather()
    
    kontroly_data = (kontroly_gather())
    kontroly_data_organized = organize_data_by_category(kontroly_data)
    




#! =================================================================================
#! ==============================Krok 2 (format)====================================
#! =================================================================================

import xml.etree.ElementTree as ET
import pandas as pd
from collections import defaultdict
from datetime import datetime
import os
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import csv

class GatherData:
    def __init__(self):
        self.special_spz = {
            "ZC 859 BR", "ZC 954 BA", "EL 638 AE", "ZH 667 CP",
            "ZC 324 BL", "ZC 685 BP", "ZC 642 BU", "ZC 651 BS",
            "ZC 137 BU", "ZC 388 BS", "ZC Z086", "ZC 206 YD",
            "kompresor", "ZC 594 BN", "VZV", "VZV2", "zeriav", "BL040EE"
        }

    def is_valid_date(self, date_string, date_format='%Y-%m-%d'):
        if not isinstance(date_string, str):
            return False
        try:
            datetime.strptime(date_string, date_format)
            return True
        except ValueError:
            return False

    def extract_data_from_excel(self, file_path):
        # Read the Excel file into a DataFrame
        try:
            df = pd.read_excel(file_path)
            print("Successfully read the Excel file.")
        
        except Exception as e:
            print(f"Error reading Excel file: {e}")
            return None

        data_dict = defaultdict(list)

        # Iterate over rows in the DataFrame to extract relevant columns
        for index, row in df.iterrows():
            column_b = row['SPZ']  # Adjust the column name as per your file
            column_e = row['Název kontr.']  # Adjust the column name as per your file
            column_j = row['Následující dne']  # Adjust the column name as per your file
            column_l = row['Násl. km od zař.']  # Adjust the column name as per your file
            column_m = row ['Akt. km od zař.'] 
            # Convert column_j to string if it's a Timestamp
            if isinstance(column_j, pd.Timestamp):
                column_j = column_j.strftime('%Y-%m-%d')  # Format it as a string

            # Determine dueTo from column J or column L if column J is None
            due_to = None
            if self.is_valid_date(column_j):
                due_to = datetime.strptime(column_j, '%Y-%m-%d')
            elif pd.notna(column_l):
                try:
                    due_to = int(column_l)  # Assume dueTo is kilometers (convert to int)
                except ValueError:
                    due_to = None

            if column_b and column_e:
                data_dict[column_b].append({
                    'kontrola': column_e,
                    'dueTo': due_to,
                    'curentKm': column_m
                    
                })
       
        return data_dict
    
    '''
    def extract_data_from_xml(self, xml_file):
        # Try different encodings if UTF-8 fails
        encodings = ['utf-8', 'utf-16', 'iso-8859-1', 'windows-1252']
        content = None
        
        for enc in encodings:
            try:
                with open(xml_file, 'r', encoding=enc) as file:
                    content = file.read()
                    print(f"Successfully read with encoding: {enc}")
                    break  # If it succeeds, break the loop
            except UnicodeDecodeError:
                print(f"Failed to decode with encoding: {enc}")
                continue
        
        if content is None:
            raise ValueError("Unable to read file with any known encoding.")
        
        # Parse the XML content
        try:
            root = ET.fromstring(content)
        except ET.ParseError as e:
            raise ValueError(f"Error parsing XML: {e}")

        # Define namespaces to be used in parsing
        namespaces = {'ss': 'urn:schemas-microsoft-com:office:spreadsheet'}
        data_dict = defaultdict(list)

        # Find the worksheet using the namespace
        worksheet = root.find('ss:Worksheet', namespaces)
        if worksheet is None:
            raise ValueError("Worksheet not found in the XML file")

        skip_header = True

        # Iterate through rows and extract data
        for table in worksheet.findall('ss:Table', namespaces):
            for row in table.findall('ss:Row', namespaces):
                if skip_header:
                    skip_header = False
                    continue

                column_b = None
                column_e = None
                column_j = None
                column_l = None

                cells = row.findall('ss:Cell', namespaces)

                for index, cell in enumerate(cells):
                    data = cell.find('ss:Data', namespaces)
                    if data is not None:
                        value = data.text
                        # Map the values to columns based on index
                        if index == 1:
                            column_b = value
                        elif index == 4:
                            column_e = value
                        elif index == 9:
                            column_j = value
                        elif index == 11:
                            column_l = value

                # Determine dueTo from column J or column L if column J is None
                due_to = None
                if column_j and self.is_valid_date(column_j.split('T')[0]):
                    due_to = datetime.strptime(column_j.split('T')[0], '%Y-%m-%d')
                elif column_l:
                    try:
                        due_to = int(column_l)  # Assume dueTo is kilometers (convert to int)
                    except ValueError:
                        due_to = None

                # Store data in dictionary if valid columns are found
                if column_b and column_e:
                    data_dict[column_b].append({
                        'kontrola': column_e,
                        'dueTo': due_to
                    })

        return data_dict
    '''

class WebDispacink:
    def __init__(self, download_dir):
        self.download_dir = download_dir
        self.driver = None

    def gather(self):
        # Configure Chrome options for downloading
        chrome_options = webdriver.ChromeOptions()
        chrome_options.add_experimental_option('prefs', {
            "download.default_directory": self.download_dir,  # Set download directory
            "download.prompt_for_download": False,  # Disable download prompt
            "download.directory_upgrade": True,  # Allow directory upgrades
            "safebrowsing.enabled": True  # Enable safe browsing
        })
        chrome_options.add_argument("--headless") 

        # Initialize the WebDriver
        self.driver = webdriver.Chrome(options=chrome_options)

        # Open the base WebDispečink URL
        self.driver.get('https://www.webdispecink.sk/sk/')

        # Wait until the "Prihlásiť" button is clickable and click it
        WebDriverWait(self.driver, 20).until(EC.element_to_be_clickable((By.ID, 'uid_login_submit')))
        login_button = self.driver.find_element(By.ID, 'uid_login_submit')
        login_button.click()

        # Wait for the login page to load, then enter the company name "MARSHALL"
        WebDriverWait(self.driver, 10).until(EC.presence_of_element_located((By.ID, 'company')))
        company_input = self.driver.find_element(By.ID, 'company')
        company_input.send_keys("MARSHALL")

        # Type "jakub" into the username field
        username_input = self.driver.find_element(By.ID, 'username')
        username_input.send_keys("jakub")

        # Type "Jakub2024*" into the password field
        password_input = self.driver.find_element(By.ID, 'password')
        password_input.send_keys("Jakub2024*")

        # Click the login button
        login_submit_button = self.driver.find_element(By.ID, 'kc-login')
        login_submit_button.click()

        # Add a short sleep to wait for the page transition
        time.sleep(1)

        # Navigate directly to the Štatistiky page
        self.driver.get('https://www.webdispecink.sk/statistika.php?zalozka=vozidla&it=2')

        # Wait for the "Export do XLSX" icon to load and click it
        WebDriverWait(self.driver, 20).until(EC.presence_of_element_located((By.CSS_SELECTOR, 'img[alt="Export do XLSX"]')))
        xlsx_export_icon = self.driver.find_element(By.CSS_SELECTOR, 'img[alt="Export do XLSX"]')
        xlsx_export_icon.click()

        # Wait for the XLSX file to download
        time.sleep(5)  # Adjust the wait time if necessary

    def formatting(self):
        # Find the latest XLSX file that starts with 'vozidla_' and ends with '.xlsx'
        latest_file = None
        for filename in os.listdir(self.download_dir):
            if filename.startswith("vozidla_") and filename.endswith(".xlsx"):
                latest_file = os.path.join(self.download_dir, filename)
                break

        if latest_file:
            # Open the downloaded XLSX file using pandas
            df = pd.read_excel(latest_file)

            # Create a dictionary from the DataFrame, mapping Column A to Column F
            result_dict = {}
            for index, row in df.iterrows():
                if index >= 9:  # Start from row 10 (index 9)
                    if row.iloc[0] == "Celkem":
                        break
                    # Remove spaces from the key
                    key = str(row.iloc[0]).replace(" ", "")
                    result_dict[key] = row.iloc[5]

            # Delete the XLSX file after processing
            os.remove(latest_file)

            return result_dict
        else:
            print("No suitable XLSX file found.")
            return None

    def close_browser(self):
        if self.driver:
            self.driver.quit()


def filter_and_categorize_controls(data, special_spz):
    control_types = {
        "l_certifikat_dict": ["dokument L- Certifikát  Lärmarmes Kraft."],
        "stk_ek_dict": ["kontrola emisná", "kontrola technická  STK"],
        "tach_dict": ["kontrola stiahnutie tachografu"],
        "dpf_dict": ["výmena DPF filtra "],
        "ciachoavnie_dict": ["kontrola pneumatik ciachovanie tachogr. "],

        'motorového_oleja': ["výmena motorového oleje "],
        'oleja_diferenciálu': ["výmena oleja diferenciálu"],
        "nastavenie_geometrie": ["servis kontrola nastavenie geometrie"],
        "kontrola_bŕzd": ["servis kontrola komplet  bŕzd"],
        "ročná_náves": ["servis ročná prehliadka náves"],
        "ročná_ťahač": ["servis ročná prehliadka ťahač"],
    }

    result_dicts = {key: {} for key in control_types.keys()}
    result_dicts["other_controls"] = {}
    result_dicts["non_trucks"] = {}  # Add non_trucks category

    # Normalize special_spz for robust matching irrespective of spaces/case
    special_spz_normalized = {str(spz).replace(" ", "").lower() for spz in special_spz}

    def format_due_to(due_to):
        if isinstance(due_to, datetime):
            return due_to.strftime('%d.%m.%Y')
        return due_to

    for vehicle, controls in data.items():
        for control in controls:
            kontrola = control['kontrola']
            matched = False

            # Always add STK/EK controls to stk_ek_dict, regardless of vehicle type
            if kontrola in control_types['stk_ek_dict']:
                if vehicle not in result_dicts['stk_ek_dict']:
                    result_dicts['stk_ek_dict'][vehicle] = []
                result_dicts['stk_ek_dict'][vehicle].append({
                    'spz': vehicle,
                    'dueTo': format_due_to(control['dueTo']),
                    'current_km': control.get('curentKm', None)
                })
                continue  # Do not add STK/EK to any other table

            # For special_spz vehicles, all non-STK/EK controls go ONLY to the second table (non_trucks / Osobné)
            normalized_vehicle = str(vehicle).replace(" ", "").lower()
            if normalized_vehicle in special_spz_normalized:
                if vehicle not in result_dicts["non_trucks"]:
                    result_dicts["non_trucks"][vehicle] = []
                result_dicts["non_trucks"][vehicle].append({
                    'spz': vehicle,
                    'dueTo': format_due_to(control['dueTo']),
                    'control_name': kontrola,
                    'current_km': control.get('curentKm', None)
                })
                continue  # Do not add to any truck-specific or first Ostatné table

            # For regular vehicles, categorize based on control type
            for dict_name, keywords in control_types.items():
                if kontrola in keywords:
                    if vehicle not in result_dicts[dict_name]:
                        result_dicts[dict_name][vehicle] = []
                    result_dicts[dict_name][vehicle].append({
                        'spz': vehicle,
                        'dueTo': format_due_to(control['dueTo']),
                        'current_km': control.get('curentKm', None)
                    })
                    matched = True
                    break

            if not matched:
                if vehicle not in result_dicts["other_controls"]:
                    result_dicts["other_controls"][vehicle] = []
                result_dicts["other_controls"][vehicle].append({
                    'spz': vehicle,
                    'dueTo': format_due_to(control['dueTo']),
                    'control_name': kontrola,
                    'current_km': control.get('curentKm', None)
                })

    return result_dicts


def transform_dict_to_list(data):
    transformed_result = {}

    for key, value in data.items():
        transformed_result[key] = []

        if isinstance(value, dict):
            unique_records = set()

            for spz, records in value.items():
                for record in records:
                    spz_value = record.get('spz', None)
                    due_to_value = record.get('dueTo', None)
                    control_name = record.get('control_name', None)
                    current_km = record.get('current_km', None)  # Get current_km

                    if spz_value is not None:
                        # Ensure current_km is included in the unique records
                        unique_records.add((spz_value, due_to_value, control_name, current_km))  

            transformed_result[key] = list(unique_records)

    return transformed_result

def format_tagra_data():
    file_path = "src/tagra.csv"
    formatted_data = {}

    with open(file_path, 'r', encoding='utf-8') as file:
        csv_reader = csv.reader(file, delimiter=';', quotechar='"')
        
        # Skip lines until we reach "0000000005."
        for row in csv_reader:
            if row[0] == "0000000005.":
                break
        
        # Process data rows
        for row in csv_reader:
            # Stop when we encounter an empty row
            if row[1] == "":
                break
            
            name = row[1]
            date_str = row[4]
            
            # Parse the date string
            date = datetime.strptime(date_str, "%d. %m. %Y").date()
            
            formatted_data[name] = date

    return formatted_data


import requests
from datetime import datetime, timezone, timedelta



def get_slovakia_time():
    api_key = "K51TZH6H2935"  # Your API key
    url = f"http://api.timezonedb.com/v2.1/get-time-zone?key={api_key}&format=json&by=zone&zone=Europe/Bratislava"
      
    try:
        response = requests.get(url)
        response.raise_for_status()
        data = response.json()
        
        # Parse the time as a datetime object
        datetime_str = data['formatted']  # Example format: '2024-10-29 15:32:00'
        slovakia_time = datetime.strptime(datetime_str, "%Y-%m-%d %H:%M:%S")
        return slovakia_time
    except requests.RequestException as e:
        print(f"Error fetching time: {e}")
        return None

def is_time_accurate(slovakia_time, tolerance_seconds=60):
    # Get the local system time
    system_time = datetime.now()
    if slovakia_time:
        time_difference = abs((system_time - slovakia_time).total_seconds())
        return time_difference <= tolerance_seconds
    return False

if __name__ == "__main__":
    print("Formating data")
    gather_data = GatherData()

    xml_file = 'src/doprava.xls'
    result_dict = gather_data.extract_data_from_excel(xml_file)

    external_time = get_slovakia_time()
    if is_time_accurate(external_time):
        download_dir = "Z:\\Jakub\\scripts\\KontrolaApp\\src" 
        web_dispacink = WebDispacink(download_dir)
        web_dispacink.gather()
        km_dict = web_dispacink.formatting()
        web_dispacink.close_browser()
    else:
        km_dict = None

    # Filter and categorize data
    categorized_data = filter_and_categorize_controls(result_dict, gather_data.special_spz)
    
    # Transform to list format
    transformed_data = transform_dict_to_list(categorized_data)
    # Main formatting process
   # tagra_data = format_tagra_data()
  
    # `transformed_data` now contains all categorized data, including special `spz` in `non_trucks`


#! =================================================================================
#! ==============================Krok 3 (Excel)=====================================
#! =================================================================================
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import NamedStyle, Font
from datetime import datetime
import re

red_font = Font(color="FF0000")  # Red color

# Create a new workbook and select the active worksheet
wb = Workbook()
ws = wb.active
ws.column_dimensions['B'].width = 17.8  
ws.column_dimensions['C'].width = 18.5  
ws.column_dimensions['D'].width = 17.5  

ws.column_dimensions['F'].width = 14.75  
ws.column_dimensions['G'].width = 18.5  
ws.column_dimensions['H'].width = 17.5  

ws.column_dimensions['J'].width = 10.4 
ws.column_dimensions['K'].width = 18.7 
ws.column_dimensions['L'].width = 37 
ws.column_dimensions['M'].width = 17.5  




# Apply text format to column C
from openpyxl.styles import NamedStyle
text_style = NamedStyle(name='text_style', number_format='@')
for row in ws.iter_rows(min_col=3, max_col=3, min_row=2, max_row=ws.max_row):
    for cell in row:
        cell.style = text_style


# Helper function to convert date strings to datetime objects for sorting
def parse_date(date_str):
    if date_str is None:
        return datetime.min  # Return a minimum date if None
    if isinstance(date_str, str):
        return datetime.strptime(date_str, "%d.%m.%Y")
    return date_str  # Return as is for non-string types like int for kilometers

def extract_kilometers(value):
    if isinstance(value, int):
        return value
    if isinstance(value, str):
        # Try to extract the leading integer
        import re
        match = re.match(r"(\d+)", value)
        if match:
            return int(match.group(1))
    return float('inf')  # Place malformed or missing values at the end


def add_data_to_sheet(start_row, start_column, end_column, headers, data_list, old_data, section_name, table_style):
    global km_dict  # Access the global km_dict
    
    # Calculate the starting column index dynamically
    column_number = ord(start_column) - ord('A') + 1
    
    # Set headers
    for col_num, header in enumerate(headers, column_number):
        cell = ws.cell(row=start_row, column=col_num)
        cell.value = header

    # Filter out entries with None or invalid 'DueTo' values, then sort if applicable
    filtered_data = [entry for entry in data_list if entry[1] is not None]
    if headers[1] == "Datum":
        sorted_data = sorted(filtered_data, key=lambda x: parse_date(x[1]) if x[1] is not None else datetime.max)
    elif headers[1] == "Kilometrov":
        sorted_data = sorted(filtered_data, key=lambda x: extract_kilometers(x[1]))
    else:
        sorted_data = data_list

    # Add data rows
    for row_index, entry in enumerate(sorted_data, start_row + 1):
        spz = entry[0]
        dueto = entry[1]
        note = None  # Default value for note

        # Modified kilometer calculation logic
        if headers[1] == "Kilometrov":
            if km_dict is not None:  # Only calculate if km_dict exists
                try:
                    # Attempt to get current_km from km_dict first
                    current_km = km_dict.get(spz.replace(" ", ""), None)
                

                    dueto_km = str(round(dueto - current_km)) if current_km is not None else "N/A"
                    
                    # Determine the appropriate string format
                    if current_km is not None and int(dueto_km) > 0:
                        dueto = f"{dueto} (do {dueto_km})"  # Format for kilometers left
                    else:
                        dueto = f"{dueto} (po {dueto_km})"  # Format for kilometers already passed
                        ws.cell(row=row_index, column=column_number + 1).font = red_font  # Change font color to red

                except (TypeError, ValueError) as e:
                    print(f"Error calculating kilometers for {spz}: {e}")
                    dueto = str(dueto)  # Just use the original value
            else:
                dueto = str(dueto)  # Just use the original value without calculation

        # Check if this entry exists in old_data and retrieve the note if it does
        for old_entry in old_data:
            if old_entry['SPZ'] == spz and old_entry['DueTo'] == dueto:
                note = old_entry['Note']
                break

        ws.cell(row=row_index, column=column_number, value=spz)
        ws.cell(row=row_index, column=column_number + 1, value=dueto)
        if note:
            ws.cell(row=row_index, column=column_number + 2, value=note)

    # Define the table range
    table_range = f"{start_column}{start_row}:{end_column}{start_row + len(sorted_data)}"
    tab = Table(displayName=section_name.replace(" ", "_"), ref=table_range)

    # Apply the specified table style
    style = TableStyleInfo(name=table_style, showFirstColumn=False,
                           showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    tab.tableStyleInfo = style
    ws.add_table(tab)

    # Return the number of rows used
    return start_row + len(sorted_data) + 2

# Functions to create each section with different table styles
def create_stk_ek_section(start_row):
    headers = ["STK + EK", "Datum", "Poznamka"]
    return add_data_to_sheet(start_row,"B", "D", headers ,transformed_data['stk_ek_dict'], kontroly_data_organized['STK + EK'], "STK_EK", "TableStyleMedium10")

def create_tach_section(start_row):
    headers = ["Stiahnutie Tach.", "Datum", "Poznamka"]
    return add_data_to_sheet(start_row,"B", "D" , headers ,transformed_data['tach_dict'], kontroly_data_organized['Tachograf'], "Tachograf", "TableStyleMedium13")

def create_dpf_section(start_row):
    headers = ["DPF čistenie", "Kilometrov", "Poznamka"]
    return add_data_to_sheet(start_row,"B", "D", headers ,transformed_data['dpf_dict'], kontroly_data_organized['DPF'], "DPF", "TableStyleMedium11")

def create_ciachovanie_section(start_row):
    headers = ["Ciachovanie", "Datum", "Poznamka"]
    return add_data_to_sheet(start_row,"B", "D", headers ,transformed_data['ciachoavnie_dict'], kontroly_data_organized['Ciachovanie'], "Ciachovanie", "TableStyleMedium9")

def create_l_certifikat_section(start_row):
    headers = ["L- Certifikát", "Datum", "Poznamka"]
    return add_data_to_sheet(start_row,"B", "D", headers ,transformed_data['l_certifikat_dict'], kontroly_data_organized['L- Certifikát'], "L_Certifikat", "TableStyleMedium11")


def motoroveho_oleja_section(start_row):
    headers = ["Motor. olej", "Kilometrov", "Poznamka"]
    return add_data_to_sheet(start_row,"F", "H", headers ,transformed_data['motorového_oleja'], kontroly_data_organized['Motor. olej'], "Motor_olej", "TableStyleMedium9")

def oleja_diferencialu_section(start_row):
    headers = ["Difer. olej", "Kilometrov", "Poznamka"]
    return add_data_to_sheet(start_row,"F", "H", headers ,transformed_data['oleja_diferenciálu'], kontroly_data_organized['Difer. olej'], "Difer_olej", "TableStyleMedium9")

def nastavenie_geometrie_section(start_row):
    headers = ["Geometria", "Datum", "Poznamka"]
    return add_data_to_sheet(start_row,"F", "H", headers ,transformed_data['nastavenie_geometrie'], kontroly_data_organized['Geometria'], "Geometria", "TableStyleMedium11")

def kontrola_brzd_section(start_row):
    headers = ["Kontrola Bŕzd", "Datum", "Poznamka"]
    return add_data_to_sheet(start_row,"F", "H", headers ,transformed_data['kontrola_bŕzd'], kontroly_data_organized['Kontrola Bŕzd'], "KontrolaBrzd", "TableStyleMedium9")

def rocna_naves_section(start_row):
    headers = ["Ročná náves", "Datum", "Poznamka"]
    return add_data_to_sheet(start_row,"F", "H", headers ,transformed_data['ročná_náves'], kontroly_data_organized['Ročná náves'], "Ročnánáves", "TableStyleMedium11")

def rocna_tahac_section(start_row):
    headers = ["Ročná tahač", "Datum", "Poznamka"]
    return add_data_to_sheet(start_row,"F", "H", headers ,transformed_data['ročná_ťahač'], kontroly_data_organized['Ročná tahač'], "Ročnátahač", "TableStyleMedium11")


def parse_date_or_int(value):
    """Helper function to parse dates or pass through integers."""
    if isinstance(value, str):
        try:
            date_value = parse_date(value)
            return date_value.timestamp()  # Convert date to timestamp for consistent comparison
        except ValueError:
            return value  # Return the value as is if not a valid date string
    elif isinstance(value, datetime):
        return value.timestamp()  # Convert datetime object to timestamp
    return value  # Return value if it's already an integer or not a date

def check_input_format(user_input):
    date_pattern = r'^\d{2}\.\d{2}\.\d{4}$'
    number_pattern = r'^\d+$'

    if re.match(date_pattern, user_input):
        return False
    elif re.match(number_pattern, user_input):
        return True
    else:
        return "Input format is unrecognized"

def others_section(start_row):
    global km_dict  # Add this line to access the global km_dict
    headers = ["Ostatné", "Datum", "Kontrola", "Poznamka"]
    
    # Using the add_data_to_sheet helper function to add data to the worksheet
    column_number = ord('J') - ord('A') + 1

    # Set headers
    for col_num, header in enumerate(headers, column_number):
        cell = ws.cell(row=start_row, column=col_num)
        cell.value = header

    # Handle data sorting and filtering
    filtered_data = [entry for entry in transformed_data['other_controls'] if entry[1] is not None]
    sorted_data = sorted(filtered_data, key=lambda x: parse_date_or_int(x[1]))

    # Add data rows
    for row_index, entry in enumerate(sorted_data, start_row + 1):
        spz = entry[0]
        dueto = entry[1]
        kontrola = entry[2]
        note = None

        # Modified kilometer calculation logic
        if check_input_format(str(dueto)):  # If dueto is in kilometer format
            if km_dict is not None:  # Only calculate if km_dict exists
                try:
                    # First try to get current_km from web interface
                    current_km = km_dict.get(spz.replace(" ", ""), None)
                    
                    # If not found in web interface, try to get it from the original data
                    if current_km is None and len(entry) > 3:
                        current_km = int(entry[3])  # Get current_km from original data
                        print(spz, "Failed to get current_km from web interface, using original data")

                    if current_km is not None:
                        dueto_km = str(round(dueto - current_km))
                        # Determine the appropriate string format
                        if int(dueto_km) > 0:
                            dueto = f"{dueto} (do {dueto_km})"
                        else:
                            dueto = f"{dueto} (po {dueto_km})"
                            ws.cell(row=row_index, column=column_number + 1).font = red_font
                    else:
                        dueto = str(dueto)  # Use original value if no current_km available

                except (TypeError, ValueError) as e:
                    print(f"Error calculating kilometers for {spz}: {e}")
                    dueto = str(dueto)  # Just use the original value
            else:
                dueto = str(dueto)  # Just use the original value without calculation

        # Check if this entry exists in old_data and retrieve the note if it does
        for old_entry in kontroly_data_organized['Ostatné']:
            if old_entry['SPZ'] == spz and old_entry['DueTo'] == dueto:
                note = old_entry['Note']
                break

        ws.cell(row=row_index, column=column_number, value=spz)   # Column J
        ws.cell(row=row_index, column=column_number + 1, value=dueto) # Column K
        ws.cell(row=row_index, column=column_number + 2, value=kontrola)
        if note:  # Only write to column M if the note is not None
            ws.cell(row=row_index, column=column_number + 3, value=note) # Column M

    # Define the table range
    table_range = f"J{start_row}:M{start_row + len(sorted_data)}"
    tab = Table(displayName="other_controls", ref=table_range)

    # Apply a different table style for distinction
    style = TableStyleInfo(name="TableStyleMedium14", showFirstColumn=False,
                           showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    tab.tableStyleInfo = style
    ws.add_table(tab)

    return start_row + len(sorted_data) + 2

def non_trucks_section(start_row):
    headers = ["Osobné", "Datum", "Kontrola", "Poznamka"]
    
    # Using the add_data_to_sheet helper function to add data to the worksheet
    column_number = ord('J') - ord('A') + 1

    # Set headers
    for col_num, header in enumerate(headers, column_number):
        cell = ws.cell(row=start_row, column=col_num)
        cell.value = header

    # Handle data sorting and filtering
    filtered_data = [entry for entry in transformed_data['non_trucks'] if entry[1] is not None]
    sorted_data = sorted(filtered_data, key=lambda x: parse_date_or_int(x[1]))

    # Add data rows
    for row_index, entry in enumerate(sorted_data, start_row + 1):
        spz = entry[0]
        dueto = entry[1]
        kontrola = entry[2]
        note = None  # Get the note directly from the data

        
        

        # Check if this entry exists in old_data and retrieve the note if it does
        for old_entry in  kontroly_data_organized['Osobné']:
            if old_entry['SPZ'] == spz and old_entry['DueTo'] == dueto:
                note = old_entry['Note']
                break  # Stop searching once a match is found

        ws.cell(row=row_index, column=column_number, value=spz)   # Column J
        ws.cell(row=row_index, column=column_number + 1, value=dueto) # Column K
        ws.cell(row=row_index, column=column_number + 2, value=kontrola)
        if note:  # Only write to column L if the note is not None
            ws.cell(row=row_index, column=column_number + 3, value=note) # Column M


    # Define the table range
    table_range = f"J{start_row}:M{start_row + len(sorted_data)}"
    tab = Table(displayName="non_trucks", ref=table_range)

    # Apply a different table style for distinction
    style = TableStyleInfo(name="TableStyleMedium12", showFirstColumn=False,
                           showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    tab.tableStyleInfo = style
    ws.add_table(tab)

def create_tagra_section(start_row):

    tagra_data = ""

    headers = ["Meno", "Datum"]
    
    # Set column widths
    ws.column_dimensions['O'].width = 19
    ws.column_dimensions['P'].width = 11.5

    # Using the add_data_to_sheet helper function to add data to the worksheet
    column_number = ord('O') - ord('A') + 1

    # Set headers
    for col_num, header in enumerate(headers, column_number):
        cell = ws.cell(row=start_row, column=col_num)
        cell.value = header

    # Add data rows
    for row_index, (name, date) in enumerate(tagra_data.items(), start_row + 1):
        ws.cell(row=row_index, column=column_number, value=name)      # Column O
        
        # Format the date as dd.mm.yyyy
        formatted_date = date.strftime("%d.%m.%Y")
        ws.cell(row=row_index, column=column_number + 1, value=formatted_date)  # Column P

    # Define the table range
    table_range = f"O{start_row}:P{start_row + len(tagra_data)}"
    tab = Table(displayName="TagraData", ref=table_range)

    # Apply the specified table style
    style = TableStyleInfo(name="TableStyleMedium13", showFirstColumn=False,
                           showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    tab.tableStyleInfo = style
    ws.add_table(tab)

    return start_row + len(tagra_data) + 2

# Call each function to create the sections
next_start_row = create_stk_ek_section(start_row=2)
next_start_row = create_tach_section(next_start_row)
next_start_row = create_dpf_section(next_start_row)
next_start_row = create_ciachovanie_section(next_start_row)
create_l_certifikat_section(next_start_row)

next_start_row = motoroveho_oleja_section(start_row=2)
next_start_row = oleja_diferencialu_section(next_start_row)
next_start_row = nastavenie_geometrie_section(next_start_row)
next_start_row = rocna_tahac_section(next_start_row)
next_start_row = rocna_naves_section(next_start_row)
kontrola_brzd_section(next_start_row)

next_start_row=others_section(start_row=2)
non_trucks_section(next_start_row+20)

#create_tagra_section(start_row=2)

# Save the workbook
wb.save("kontroly.xlsx")

print("Excel file created successfully!")


#==============================================

# Firebase Firestore data upload from vozidla.xls
import firebase_admin
from firebase_admin import credentials, firestore
import pandas as pd
from datetime import datetime

def upload_vozidla_data_to_firestore():
    try:
        # Path to your service account key
        cred = credentials.Certificate("firebase-service-account.json")
        # Prevent double initialization if script is re-run interactively
        if not firebase_admin._apps:
            firebase_admin.initialize_app(cred)
        db = firestore.client()

        # Read vozidla.xls file
        vozidla_file = "src/vozidla.xls"
        df = pd.read_excel(vozidla_file)
        
        # Extract data from columns B (name) and S (kilometers)
        # Note: pandas uses 0-based indexing, so column B is index 1, column S is index 18
        vozidla_data = []
        
        for index, row in df.iterrows():
            name = row.iloc[1]  # Column B (index 1)
            kilometers = row.iloc[18]  # Column S (index 18)
            
            # Skip rows where name is empty or NaN
            if pd.notna(name) and str(name).strip() != "":
                vozidla_data.append({
                    "name": str(name).strip(),
                    "kilometers": int(kilometers) if pd.notna(kilometers) else 0,
                    "timestamp": datetime.now()
                })
        
        # Clear existing data in the vehicles_km collection
        vehicles_collection = db.collection("vehicles_km")
        existing_docs = vehicles_collection.stream()
        for doc in existing_docs:
            doc.reference.delete()
        
        # Upload new data using SPZ as document ID
        for vehicle in vozidla_data:
            spz = vehicle["name"].replace(" ", "")  # Use the SPZ as document ID and remove spaces
            # Remove the name field since it's now the document ID
            vehicle_data = {
                "kilometers": vehicle["kilometers"],
                "timestamp": vehicle["timestamp"]
            }
            vehicles_collection.document(spz).set(vehicle_data)
        
        print(f"Successfully uploaded {len(vozidla_data)} vehicle records to Firebase!")
        return True
        
    except Exception as e:
        print(f"Error uploading vozidla data to Firebase: {e}")
        return False

# Automatically upload vozidla data to Firestore after Excel is created
upload_vozidla_data_to_firestore()

input("Press Enter to close.")